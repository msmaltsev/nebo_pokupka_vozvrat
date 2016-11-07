# -*- coding: cp1251 -*-

import xlrd, re, os, string, pprint, string, codecs
import openpyxl as op

def saveIntoWb(tables): #type=dict
    wb = op.Workbook()
    for i in tables.keys():
        ws = wb.create_sheet(0)
        ws.title = i
        data = tables[i]
        for nrow in range(0, len(data)):
            res_nrow = nrow + 1
            row_data = data[nrow]
            for ncol in range(0, len(row_data)):
                ws.cell(row=res_nrow, column=ncol + 1).value = row_data[ncol]

    wb.save('result.xlsx')


def colIndexes(caps=False):
    a = list(string.ascii_lowercase)
    if caps:
        b = [i.upper() for i in a]
        a = b
    expansion = []
    for i in a:
        for k in a:
            expansion.append(i + k)
    a += expansion
    d = {a[i]:i for i in range(len(a))}
    return d


def getTable(fname):
    sb = xlrd.open_workbook(fname, logfile=open(os.devnull, 'w'))
    ws = sb.sheet_by_index(0)
    rows = ws.nrows
    cols = ws.ncols
    result = []
    row = 0
    while row < rows:
        result.append([ws.cell_value(row, i) for i in range(0, cols)])
        row += 1
    return result


def extractTable(table, cols, cond_col, condition): #type(table)=list, type(cols)=str, type(condition)=str
    ci = colIndexes()
    result = []
    head = []
    i = table[0]
    for c in cols:
        head.append(i[ci[c]])
    result.append(head+[u'strich_code',u'articul'])
    
    for i in table:
        if i[ci[cond_col]] == condition:
            new_row = []
            for c in cols:
                new_row.append(i[ci[c]])

            new_row_ = []
            for i in new_row:
                try:
                    s = int(i)
                    new_row_.append(s)
                except:
                    new_row_.append(i)
            new_row = new_row_
            result.append(new_row)
        else:
            pass
    return result

def removePunctuation(line):
    exclude = string.punctuation
    exclude_ = exclude.replace('-','')
    exclude = exclude_.replace('_','')
    exclude += ' '
    s = ''.join(ch for ch in line if ch not in exclude)
    return s

def getStrichCode(table): #type=list
    newtab = []
    for row in table:
        m = re.search(u'(шк(.+))', row[-1])
        n = re.search(u'(арт([^ш]+))', row[-1])
        if m is not None:
            strich = removePunctuation(m.group(2))
            row.append(strich)
        else:
            row.append('')
        if n is not None:
            articul = removePunctuation(n.group(2))
            row.append(articul)
        else:
            row.append('')
        newtab.append(row)
    return newtab

def xmlifySheet(table): # type = list    
    result = u'<xml>\r\n'
    l = table[1:]
    for i in l:
        result += u'\t<stock>\r\n'
        result += u'\t\t<shipment_number>%s</shipment_number>\r\n'%i[0]
        result += u'\t\t<cost_per_one>%s</cost_per_one>\r\n'%i[1]
        result += u'\t\t<money_received>%s</money_received>\r\n'%i[2]
        result += u'\t\t<status>%s</status>\r\n'%i[3]
        result += u'\t\t<status_date>%s</status_date>\r\n'%i[4]
        result += u'\t\t<stock_name>%s</stock_name>\r\n'%i[5]
        result += u'\t\t<strich_code>%s</strich_code>\r\n'%i[6]
        result += u'\t\t<articul>%s</articul>\r\n'%i[7]
        result += u'\t</stock>\r\n'
    result += u'</xml>\r\n'
    return result
        

if __name__ == '__main__':
    table_arr = getTable('source.xls')
    d = {}
    d['pokupka'] = getStrichCode(extractTable(table_arr, 'aghlmr', 'l', u'Вручен'))
    d['vozvrat'] = extractTable(table_arr, 'aghlmrpqt', 'p', u'Возврат')
    saveIntoWb(d)
    xml = xmlifySheet(d['pokupka'])
    f = codecs.open('pokupka_xml.xml', 'w', 'utf8')
    f.write(xml)
    f.close()
    
    
    
